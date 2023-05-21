"""
schi15.py

Preprocess Schi15 data (data from Schiemann et al 2015, Cell Reports paper)

Contributors: salvadordura@gmail.com
"""

import numpy as np
from scipy.io import loadmat
import os 
import pandas
import openpyxl as xl
from matplotlib import pyplot as plt


def getMovementVsQuietPeriods(motionIndex, motionIndexTimes, threshold = 0.5, minDur = 2.0, bufferDur = 0.5, exclude='on'):
    '''
    Classify movement periods based on motion indexs
    '''

    # find quiet vs movement periods
    lastSwitchOn = 0
    lastSwitchOff = 0

    if motionIndex[0] <= threshold:
        state = 0
    else:
        state = 1

    onPeriods = []
    offPeriods = []

    for i,(m, t) in enumerate(zip(motionIndex, motionIndexTimes)):
        if m > threshold and state == 0:
            lastSwitchOn = t
            state = 1
            offPeriods.append([float(lastSwitchOff), float(t)])

        elif m <= threshold and state == 1:
            lastSwitchOff = t
            state = 0
            onPeriods.append([float(lastSwitchOn), float(t)])

    movementPeriods = [x for x in onPeriods if x[1] - x[0] >= minDur]
    quietPeriods = offPeriods
    if exclude == 'on':  # exclude 0.5s around all on periods (threshold crossings)
        excludeQuietPeriods = [[x[0] - 0.5, x[0]] for x in onPeriods] + [[x[1], x[1] + 0.5] for x in onPeriods]
    elif exclude == 'movement':  # exclude 0.5s only around movement periods (>2s duration)
        excludeQuietPeriods = [[x[0] - 0.5, x[0]] for x in movementPeriods] + [[x[1], x[1] + 0.5] for x in movementPeriods]

    # exclude buffer periods (0.5s before and after movement) from quiet periods
    for i,quietPeriod in enumerate(quietPeriods):
        for excludeQuietPeriod in excludeQuietPeriods:
            if excludeQuietPeriod[0] <= quietPeriod[0] < excludeQuietPeriod[1] and \
                excludeQuietPeriod[0] <= quietPeriod[1] < excludeQuietPeriod[1]: # if start and end in exclude
                quietPeriods[i] = [-1,-1]  # remove
            elif excludeQuietPeriod[0] <= quietPeriod[0] < excludeQuietPeriod[1] and \
                quietPeriod[1] > excludeQuietPeriod[1]: # if start in exclude and end outside
                quietPeriods[i][0] = excludeQuietPeriod[1]  # set new end to exclude
            elif excludeQuietPeriod[0] <= quietPeriod[1] < excludeQuietPeriod[1] and \
                quietPeriod[0] < excludeQuietPeriod[0]: # if end in exclude and start before
                quietPeriods[i][1] = excludeQuietPeriod[0]  # set new end to exclude

    return movementPeriods, quietPeriods



def readSpikeTimesAndExcelMetadata():
    '''
    Read experimental data and store in pandas data fram
    '''

    threshold = 0.5
    minDur = 2.0
    bufferDur = 0.5

    # load data
    rootFolder = '../data/Schi15/'
    dataFolders = {'main': 'Main_Dataset_Extracted/', 'mth-inact': 'MThInact_Dataset_Extracted/', 'na-block': 'Na_Dataset_Extracted/'}
    metadataFiles = {'main': 'Main_Dataset_AdditionalCellIInfo_meanFR.xlsx', 'mth-inact': 'MThInact_Dataset_AdditionalCellIInfo_meanFR.xlsx', 'na-block': 'NA_Dataset_AdditionalCellIInfo_meanFR.xlsx'}
    metadataSheets = {'main': 'Main Dataset_Cell-Info_cellN', 'mth-inact': 'MThInact Dataset_CellInfo_cellN', 'na-block': 'Schiemann_CellReport_2015_FR1.t'}
    metadataFields = {'cell': 1, 'depth': 4, 'code': 5, 'type': 6} 
    metadataStartRow = {'main': 2, 'mth-inact': 3, 'na-block': 3}

    '''
    code: 51 - L5enh; 52 - L5supp; 23 - L2/3; 56 - L5 na-block; 55 - L5 MTh-Inact
    '''

    df = pandas.DataFrame()

    for condition, dataFolder in dataFolders.items():
        cells = [f[:-4] for f in os.listdir(rootFolder + dataFolder) if os.path.isfile(os.path.join(rootFolder + dataFolder, f)) and f.endswith('.mat')]
        cellsData = {}

        for cell in cells:
            data = loadmat(rootFolder+dataFolder+cell+'.mat')

            # read variables
            spikeTimes = list(data['Data'][0][0][0][0])
            motionIndex = data['Data'][0][0][1]
            motionIndexTimes = data['Data'][0][0][2]

            # classify periods based on motion index
            movePeriods, quietPeriods = getMovementVsQuietPeriods(motionIndex, motionIndexTimes, threshold, minDur, bufferDur)

            # assign spikes to different 
            spikeTimes = [spk for spk in spikeTimes if spk >= 0.0]  # get rid of negative spiket times
            moveSpikes = []
            quietSpikes = []
            for spk in spikeTimes:
                for movePeriod in movePeriods:
                    if movePeriod[0] <= spk < movePeriod[1]:
                        moveSpikes.append(spk)
                for quietPeriod in quietPeriods:
                    if quietPeriod[0] <= spk < quietPeriod[1]:
                        quietSpikes.append(spk)
            
            # calculate firing rates
            moveTime = np.sum([x[1] - x[0] for x in movePeriods])
            quietTime = np.sum([x[1] - x[0] for x in quietPeriods])  

            moveFR = len(moveSpikes) / moveTime if moveTime > 0 else 0
            quietFR = len(quietSpikes) / quietTime if quietTime > 0 else 0

            cellsData[int(cell.split('Cell')[1])] = {'condition': condition, 'quietFR': quietFR, 'moveFR': moveFR, 'quietSpikes': quietSpikes, 'moveSpikes': moveSpikes} # , 'quietSpikes': quietSpikes, 'moveSpikes': moveSpikes}
        
        # read data from excel
        wb = xl.load_workbook(rootFolder+metadataFiles[condition])
        sheet = wb[metadataSheets[condition]]
        numRows = sheet.max_row

        for row in range(metadataStartRow[condition], numRows + 1):
            try:
                cellid = int(sheet.cell(row=row, column=1).value)
            except:
                cellid = -1
            if cellid in cellsData:
                cellsData[sheet.cell(row=row, column=1).value].update({k: sheet.cell(row=row, column=v).value for k, v in metadataFields.items()})

        if df.empty:
            df = pandas.DataFrame(list(cellsData.values()))
        else:
            df = pandas.concat([df, pandas.DataFrame(list(cellsData.values()))])

        # example of querying: df.query('condition=="main" and type=="IT"') 
    
    return df


def readExcelFiringRatesAndMetada():
    '''
    Read experimental data and store in pandas data fram
    '''

    # load data
    rootFolder = '../data/Schi15/'
    metadataFiles = {'main': 'Main_Dataset_AdditionalCellIInfo_meanFR.xlsx', 'mth-inact': 'MThInact_Dataset_AdditionalCellIInfo_meanFR.xlsx', 'na-block': 'NA_Dataset_AdditionalCellIInfo_meanFR.xlsx'}
    metadataSheets = {'main': 'Main Dataset_Cell Info_groups', 'mth-inact': 'MThInact Dataset_CellInfo_cellN', 'na-block': 'Schiemann_CellReport_2015_FR1.t'}
    metadataFields = {'cell': 1, 'depth': 4, 'code': 5, 'cell_class': 6, 'quietFR': 8, 'moveFR': 9, 'type': 10} 
    metadataStartRow = {'main': 2, 'mth-inact': 3, 'na-block': 3}

    '''
    code: 51 - L5enh; 52 - L5supp; 23 - L2/3; 56 - L5 na-block; 55 - L5 MTh-Inact
    '''

    df = pandas.DataFrame()

    for condition in metadataFiles.keys():

        # read data from excel
        wb = xl.load_workbook(rootFolder+metadataFiles[condition])
        sheet = wb[metadataSheets[condition]]
        numRows = sheet.max_row

        cellsData = {}

        for row in range(metadataStartRow[condition], numRows + 1):
            try:
                cellid = int(sheet.cell(row=row, column=1).value)
                cellsData[cellid] = {k: sheet.cell(row=row, column=v).value for k, v in metadataFields.items()}
                cellsData[cellid]['condition'] = condition
            except:
               pass
               #print('Skipping row: %s' % (row))

        if df.empty:
            df = pandas.DataFrame(list(cellsData.values()))
        else:
            df = pandas.concat([df, pandas.DataFrame(list(cellsData.values()))])    
    
    return df


def getIhVsVLRates(df, avg = 'mean'):
    
    ihValues = ['low', 'medium', 'high']
    VLValues = ['low', 'medium', 'high']

    rates = np.zeros((len(ihValues), len(VLValues)))

    if avg == 'mean':
        # ih=medium; VL=medium: quiet (wakefulness) 
        rates[1, 1] = df[df.condition=="main"].quietFR.mean() 

        # ih=low, VL=high: movement (self-paced, voluntary) 
        rates[0, 2] = df[df.condition=="main"].moveFR.mean() #

        # ih=medium; VL=low: inactive VL; quiet 
        rates[1, 0] = df[df.condition=="mth-inact"].quietFR.mean()

        # ih=low; VL=low: inactive VL; movement -->  ih low (0.25); only bkg inputs; VL = 0 (~2Hz); 
        # from data = 4.7Hz (!)
        rates[0, 0] = df[df.condition=="mth-inact"].moveFR.mean()
        
        # ih=high; VL=medium: NA-R antagonist; quiet --> ih high (?); only bkg inputs0 (~1.8Hz); 
        # from data = 1.5Hz (exclude='movement'), 1.3Hz (exclude='on')
        rates[2, 1] = df[df.condition=="na-block"].quietFR.mean()
        
        # ih=high; VL=high: NA - R antagonist; movement - ->ih high(?); bkg inputs + high VL(bimodal: ~ 1 HZ vs ~ 7 Hz; ~1.4Hz); 
        # from data = 1.3Hz
        rates[2, 2] = df[df.condition=="na-block"].moveFR.mean()
    
    elif avg == 'median':
        rates[1, 1] = df[df.condition=="main"].quietFR.median() 
        rates[0, 2] = df[df.condition=="main"].moveFR.median() 
        rates[1, 0] = df[df.condition=="mth-inact"].quietFR.median()
        rates[0, 0] = df[df.condition=="mth-inact"].moveFR.median()
        rates[2, 1] = df[df.condition=="na-block"].quietFR.median()
        rates[2, 2] = df[df.condition=="na-block"].moveFR.median()
    return rates


def addScaleBar(timeRange=[0, 1000], loc=1):
    from netpyne.support.scalebar import add_scalebar
    ax = plt.gca()
    sizex = (timeRange[1]-timeRange[0])/20.0
    #yl = plt.ylim()
    #plt.ylim(yl[0]-0.2*(yl[1]-yl[0]), yl[1])
    add_scalebar(ax, hidex=False, hidey=True, matchx=False, matchy=True, sizex=sizex, sizey=None, unitsx='ms', unitsy='mV', scalex=10., scaley=1, loc=loc, pad=-1, borderpad=0.5, sep=4, prop=None, barcolor="black", barwidth=3)
    plt.axis('off')


def readTraces(rootFolder, inputFolder, dataFile):
    from scipy.io import loadmat
    
    dataFileFullPath = rootFolder + inputFolder + dataFile

    data = loadmat(dataFileFullPath, simplify_cells=1)
    dataV = data['all_var']['basic']['trace']  

    try:
        onlimits = [float(x) for x in data['all_var']['movtrig']['onlimitbigbin']]
    except:
        onlimits = [float(data['all_var']['movtrig']['onlimitbigbin'])]

    try:
        offlimits = [float(x) for x in data['all_var']['movtrig']['offlimitbigbin']]
    except:
        offlimits = [float(data['all_var']['movtrig']['offlimitbigbin'])]

        
    dt = float(data['all_var']['basic']['dt'])
    fs = 1 / dt

    if onlimits[0] < offlimits[0]:
        timeRanges = [[on, off] for on, off in zip(onlimits, offlimits)] \
                    + [[off, on] for on, off in zip(onlimits[1:], offlimits)] \
                    + [[0, onlimits[0]]] \
                    + [[offlimits[-1], len(dataV)]]
    
        timeLabels = ['move' if t[0] in onlimits else 'quiet' for t in timeRanges]  # list of boolean indicating if move period
        timeRanges = [[t[0] * 1000 * dt, t[1] * 1000 * dt] for t in timeRanges]

    timeRanges.append([0, timeRanges[-1][1]])  # add full duration
    timeLabels.append('quietmove')

    return dataV, fs, timeRanges, timeLabels, onlimits, offlimits 


def plotTracesMain():
    rootFolder = '../data/Schi15/'
    inputFolder = '/Main_Dataset_Movement_Times/' #Additional_Files_July_2020/'
    outputFolder = '/voltageTraces/'

    state = 'quiet' # 'move''

    if state == 'quiet':
        # for quiet (datafile 228 - IT)
        dataFile = '228.mat' 
        dataV, fs, timeRanges, timeLabels, onlimits, offlimits = readTraces(rootFolder, inputFolder, dataFile )
        dt = 1 / fs
        quietTime = (offlimits[-2] * 1000 * dt)
        print(quietTime)
        timeRanges = [[quietTime, quietTime+2000]]
        timeLabels = ['quiet'] 
        figSize = (15*1.5*0.5, 2.5)
        yamp = 60

    elif state == 'move':

        # for move (dataFile 223 - PT)
        dataFile = '223.mat' 
        dataV, fs, timeRanges, timeLabels, onlimits, offlimits = readTraces(rootFolder, inputFolder, dataFile )
        dt = 1 / fs
        moveTime = (onlimits[-2] * 1000 * dt)
        print(moveTime)
        timeRanges = [[moveTime - 1000, moveTime + 5000]]
        timeLabels = ['quietmovequiet'] 
        figSize = (15*1.5, 2.5)
        yamp = 60

    fontSize = 20
    
    color = 'black'
    axis = False
    ylim = [-60, 5]

    plt.rcParams.update({'font.size': fontSize})
    fontsiz = fontSize
    
    for timeRange, timeLabel in zip(timeRanges, timeLabels):
        v = dataV[int(timeRange[0] * fs / 1000.):int(timeRange[1]*fs/1000.)]
        t = range(len(v))

        
        plt.figure(figsize=figSize)
        plt.plot(t, v, linewidth=1, color=color, label='L5B Experiment')
        baseline = min(v)
        #addScaleBar()
        plt.axis('off')
        plt.ylim(baseline, baseline+yamp)
        ax = plt.gca()
        plt.title('')
         
        if timeLabel == 'quietmove': # add dashed lines
            for onlimit in onlimits:
                y = range(ylim[0], ylim[1])
                x = [onlimit ] * len(y)
                plt.plot(x, y, 'g--')
            for offlimit in offlimits:
                y = range(ylim[0], ylim[1])
                x = [offlimit] * len(y)
                plt.plot(x, y, 'r--')
        
        if timeLabel == 'quietmovequiet':  # add dashed lines
            offset = 10000
            for onlimit in [offset, onlimits[-1] - onlimits[-2] + offset]:
                y = range(ylim[0], ylim[1])
                x = [onlimit] * len(y)
                plt.plot(x, y, 'g--')
            for offlimit in [offlimits[-2] - onlimits[-2] + offset]:
                y = range(ylim[0], ylim[1])
                x = [offlimit] * len(y)
                plt.plot(x, y, 'r--')

        plt.ylim(baseline, baseline+yamp)
        plt.tight_layout()

        plt.savefig('%s/%s/%s_%s_%d_%d_main_truncated.png' % (rootFolder, outputFolder, dataFile[:-4], timeLabel, timeRange[0], timeRange[1]), dpi=200)


def plotTracesMThInact():
    rootFolder = '../data/Schi15/'
    inputFolder = '/MthInact_Dataset_Movement_Times/' #Additional_Files_July_2020/'
    outputFolder = '/voltageTraces/'
    dataFiles = ['5002.mat'] #['%d.mat' % (x) for x in range(5000, 5006)]

    for dataFile in dataFiles:
        print(dataFile)

        dataV, fs, timeRanges, timeLabels, onlimits, offlimits = readTraces(rootFolder, inputFolder, dataFile )

        # timeRanges = [timeRanges[-1]]
        # timeLabels = [timeLabels[-1]]

        moveTime = (onlimits[1] * 1000 / fs)
        print(moveTime)
        timeRanges = [[moveTime - 1000, moveTime + 2000]]
        timeLabels = ['quietmovequiet'] 

        fontSize = 20
        
        color = 'black'
        axis = False
        ylim = [-60, 5]

        plt.rcParams.update({'font.size': fontSize})
        fontsiz = fontSize
    
        for timeRange, timeLabel in zip(timeRanges, timeLabels):
            v = dataV[int(timeRange[0] * fs / 1000.):int(timeRange[1]*fs/1000.)]
            t = range(len(v))

            figSize=(min(30, 2 * (timeRange[1] - timeRange[0]) / 1000.0), 3.5)
            figSize = (15*1.5, 3.5)
            plt.figure(figsize=figSize)
            plt.plot(t, v, linewidth=1.5, color=color, label='L5B Experiment')
            plt.xlabel('Time (ms)', fontsize=fontsiz)
            plt.ylabel('mV', fontsize=fontsiz)
            #plt.xlim(t) #timeRange)
            #if ylim: plt.ylim(ylim)
            plt.axis('off')
            #addScaleBar(timeRange) # same scale as model

            if timeLabel == 'quietmove': # add dashed lines
                for onlimit in onlimits:
                    y = range(ylim[0], ylim[1])
                    x = [onlimit ] * len(y)
                    plt.plot(x, y, 'g--')
                for offlimit in offlimits:
                    y = range(ylim[0], ylim[1])
                    x = [offlimit] * len(y)
                    plt.plot(x, y, 'r--')
            
            if timeLabel == 'quietmovequiet':  # add dashed lines
                offset = 10000
                for onlimit in [offset]:
                    y = range(ylim[0], ylim[1])
                    x = [onlimit] * len(y)
                    plt.plot(x, y, 'g--')
                # for offlimit in [offlimits[1] - onlimits[1] + offset]:
                #     y = range(ylim[0], ylim[1])
                #     x = [offlimit] * len(y)
                #     plt.plot(x, y, 'r--')

            plt.savefig('%s/%s/%s_%s_%d_%d_mth.png' % (rootFolder, outputFolder, dataFile[:-4], timeLabel, timeRange[0], timeRange[1]), dpi=300)




# ------------------------------------------------------------------------
# Main code
# ------------------------------------------------------------------------
if __name__ == '__main__':
    df = readExcelFiringRatesAndMetada()
    df.dropna()
    ratesExcel = getIhVsVLRates(df)
    plotTracesMain()
    plotTracesMThInact()


